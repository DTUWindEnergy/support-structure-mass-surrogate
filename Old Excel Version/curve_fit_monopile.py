# -*- coding: utf-8 -*-
"""
Created on Mon Jan 10 11:00:22 2022

@author: mikf
"""

from scipy.optimize import curve_fit
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import openturns as ot
from sklearn.preprocessing import MinMaxScaler
import openpyxl


plt.close('all')
'''
['IP',
 'RP',
 'SP',
 'set_id',
 'D',
 'HHub_Ratio',
 'HHub',
 'HTrans',
 'PileDepth',
 'WaterDepth',
 'WaveHeight',
 'WavePeriod',
 'WindSpeed',
 'monopile_mass',
 'tower_mass',
 'total_mass']
'''
data = pd.read_csv('data/tower_mass_results.dat', sep=' ', )
df = data[data.columns[:-1]]
df.columns = data.columns[1:]
in_cols = ['RP', 'D', 'SP', 'HTrans', 'HHub_Ratio',
           'WaterDepth', 'WaveHeight', 'WavePeriod', 'WindSpeed']#, 'IP'
long_in = ['Rated Power', 'Rotor Diameter', 'Specific Power', 'Transition Piece Height',
            'Hub Height Ratio', 'Water Depth', 'Wave Height', 'Wave Period', 'Wind Speed']
short_in = ['RP', 'D', 'SP', 'TPH',
            'HHR', 'WD', 'WH', 'WP', 'WS']
conv_dict = {k: v for k, v in zip(in_cols, short_in)}
out_cols = ['monopile_mass', 'tower_mass', 'total_mass']
long_out = ['Monopile Mass', 'Tower Mass', 'Total Mass']
short_out = ['Monopile Mass', 'Tower Mass', 'Total Mass']
conv_dict_out = {k: v for k, v in zip(out_cols, short_out)}
conv_dict.update(conv_dict_out)
outliers = df.iloc[[ 3884,  3901,  4337,  4374,  4650,  7678,  7851,  7862, 10172,
       10999, 12231, 13036, 13544, 14696, 16594, 17184, 24249, 25087,
       25525, 27408, 29344, 29650, 29669, 29783, 29830, 32033, 35132,
       37751, 38275, 38892, 38996, 39924, 41520, 48054, 53909, 54198,
       54327, 54884, 55136, 55439, 55476, 55573, 55814, 55965, 57194,
       57344, 58064, 58110, 66574, 77821, 83776, 85725, 87207, 93091,
       93779]].index
df.drop(outliers, inplace=True)
initial_powers = [3.4, 10.0, 15.0]
IP = initial_powers[2]
df = df[df.IP==IP]
# set_id = initial_powers[0]
df = df[df.set_id==1]
df.reset_index(drop=True, inplace=True)
inp = df[in_cols]
out = df[out_cols]
count = out.shape[0]
update_excel = False

def get_r2(data, prediction):
    residuals = data - prediction
    ss_res = np.sum(residuals**2)
    ss_tot = np.sum((data-np.mean(data))**2) 
    r2_power = 1 - (ss_res / ss_tot)
    return r2_power
name_map = conv_dict#{x:x for x in list(df)}

def train_model(df):

    input_db = df[in_cols]
    output_db = df[out_cols]
   
    # # Input and output names.
    input_channel_names = input_db.columns.to_list()
    output_channel_names = output_db.columns.to_list()
    
    # Numpy versions of the input and output dataset.
    input_dataset = input_db.to_numpy()
    output_dataset = output_db.to_numpy()
    n_output = output_dataset.shape[1]
    
    # # %% Center and scale the input and output dataset.
    
    # Center and scale the input dataset.
    input_scaler = MinMaxScaler(feature_range=(-1.0, 1.0))
    input_dataset_scaled = input_scaler.fit_transform(input_dataset)
    
    # Center and scale the output dataset.
    output_scalers = {}
    output_dataset_scaled = np.empty_like(output_dataset)
    for i in range(len(output_channel_names)):
        output_channel_name = output_channel_names[i]
        output_scalers[output_channel_name] = MinMaxScaler(
                feature_range=(-0.7, 0.7))
        output_dataset_scaled[:, i] = \
            output_scalers[output_channel_name].fit_transform(
                output_dataset[:, [i]]).ravel()
    
    
    # Fit the model
    # Compose the names for the linear and quadratic dependencies.
    predicted_output = pd.DataFrame(columns=output_channel_names)
    names = []
    for i in range(len(input_channel_names)):
        for j in range(0, i+1):
            names.append(name_map[input_channel_names[i]] + ' * ' + name_map[input_channel_names[j]])
    dependencies = pd.DataFrame(
        index=[name_map[input_channel_name] for input_channel_name in input_channel_names] + names,
        columns=[name_map[output_channel_name] for output_channel_name in output_channel_names])
    
    def predict(inputs, constant, linear, quadratic):
        out = np.zeros(inputs.shape[1])
        for i, inp in enumerate(inputs):
            out += inp * linear[i]
            for j, jnp in enumerate(inputs):
                out += 0.5 * np.multiply(inp, jnp) * quadratic[i, j]
        out += constant
        return out
    
    models = []
    coefficients = {}
    for i_output_channel in range(n_output):
        model = ot.QuadraticLeastSquares(input_dataset_scaled, output_dataset_scaled[:, [i_output_channel]])
        model.run()
        models.append(model)
        
        # Get linear and quadratic dependencies of output from input variables.
        constant = np.squeeze(np.array(model.getConstant()))
        linear = np.squeeze(np.array(model.getLinear()))
        quadratic_full = np.squeeze(np.array(model.getQuadratic()))
        coefficients[output_channel_names[i_output_channel]] = {'constant': constant,
                                                                'linear': linear,
                                                                'quadratic': quadratic_full,
                                                                }
        
        quadratic = quadratic_full - np.diag(np.diag(quadratic_full) * 0.5)
        quadratic = quadratic[np.tril_indices_from(quadratic)]
        dependencies.iloc[:, i_output_channel] = np.concatenate((linear.ravel(), quadratic))
        output_channel_name = output_channel_names[i_output_channel]
        responseSurface = model.getMetaModel()
        scaled_output = responseSurface(input_dataset_scaled)
        out = output_scalers[output_channel_name].inverse_transform(scaled_output).ravel()
        predicted_output[output_channel_name] = out
        df.loc[:, output_channel_name + '_fit'] = out
        df.loc[:, output_channel_name + '_scaled'] = scaled_output
    return input_scaler, output_scalers, df, dependencies, models, coefficients, predicted_output

input_scaler, output_scalers, df, dependencies, models, coefficients, predicted_output = train_model(df)

in_min = df[in_cols].to_numpy().min(axis=0)
in_max = df[in_cols].to_numpy().max(axis=0)
out_min = df[out_cols].to_numpy().min(axis=0)
out_max = df[out_cols].to_numpy().max(axis=0)

if update_excel:
    wb = openpyxl.load_workbook('surrogate/monopile_surrogate.xlsx')
    
    out_key = 'monopile_mass'
    out_key_no = 0
    sheet = wb[f'{out_key}_IP_{int(IP)}']
    for n, v in enumerate(in_min):
        sheet.cell(row=17, column=2+n, value=v)
    for n, v in enumerate(in_max):
        sheet.cell(row=18, column=2+n, value=v)
    sheet['B19'] = out_min[out_key_no]
    sheet['B20'] = out_max[out_key_no]
    sheet['K22'] = float(coefficients[out_key]['constant'])
    for n, v in enumerate(coefficients[out_key]['linear']):
        sheet.cell(row=6, column=n+2, value=v)
    for i, vs in enumerate(coefficients[out_key]['quadratic']):
        for j, v in enumerate(vs):
            sheet.cell(row=8+i, column=2+j, value=v)
    
    out_key = 'tower_mass'
    out_key_no = 1
    sheet = wb[f'{out_key}_IP_{int(IP)}']
    for n, v in enumerate(in_min):
        sheet.cell(row=17, column=2+n, value=v)
    for n, v in enumerate(in_max):
        sheet.cell(row=18, column=2+n, value=v)
    sheet['B19'] = out_min[out_key_no]
    sheet['B20'] = out_max[out_key_no]
    sheet['K22'] = float(coefficients[out_key]['constant'])
    for n, v in enumerate(coefficients[out_key]['linear']):
        sheet.cell(row=6, column=n+2, value=v)
    for i, vs in enumerate(coefficients[out_key]['quadratic']):
        for j, v in enumerate(vs):
            sheet.cell(row=8+i, column=2+j, value=v)
    
    
    wb.save('surrogate/monopile_surrogate.xlsx')


# Plot mass vs rated power
plt.figure() 
rps = sorted(df.RP.unique())
sps = sorted(df.SP.unique())
colors = np.array([
    ['orange', 'darkorange'],
    ['lime', 'darkgreen'],
    ['blue', 'darkblue',],
    ['red', 'darkred'],
         ]).ravel()
for i, sp in enumerate(sps):
    this_df = df.copy()
    this_df = this_df[this_df.SP == sp]
    y_data = []
    y_fit = []
    for rp in rps:
        y_data.append(this_df[this_df.RP == rp].total_mass.mean())
        y_fit.append(this_df[this_df.RP == rp].total_mass_fit.mean())
        
    line = plt.plot(rps, y_data, label=f'sp:{sp:.0f} W/m2', linestyle='-')
    plt.plot(rps, y_fit, linestyle='--', color=line[0].get_color())
plt.ylim([0, plt.ylim()[1]])
plt.legend()
plt.grid()
plt.xlabel('Rated power [MW]')
plt.ylabel('Total mass [kg]')
plt.savefig('mass_vs_rp')



plt.figure() 
wds = sorted(df.WaterDepth.unique())[::4]
for i, wd in enumerate(wds):
    this_df = df.copy()
    this_df = this_df[this_df.WaterDepth == wd]
    d_bins = pd.cut(this_df.D, 10)
    this_df['d_bins'] = d_bins
    d_bins_uni = sorted(d_bins.unique())
    xs = [x.mid for x in d_bins_uni]
    y_data = []
    y_fit = []
    for d_bin in d_bins_uni:
        y_data.append(this_df[this_df['d_bins'] == d_bin].total_mass.mean())
        y_fit.append(this_df[this_df['d_bins'] == d_bin].total_mass_fit.mean())
        
    line = plt.plot(xs, y_data, label=f'wd:{wd:.0f} m', linestyle='-')
    plt.plot(xs, y_fit, linestyle='--', color=line[0].get_color())
plt.ylim([0, plt.ylim()[1]])
plt.legend()
plt.grid()
plt.xlabel('Rotor diameter [m]')
plt.ylabel('Total mass [kg]')
plt.savefig('mass_vs_d')


indx = np.flip(np.argsort(np.abs(dependencies['Total Mass'])))
ax = dependencies.iloc[indx[:20]].plot.bar()

textstr = '\n'.join([f'{k}: {v}' for k, v in {k: v for k, v in zip(short_in, long_in)}.items()])

# ax.hist(x, 50)
# these are matplotlib.patch.Patch properties
props = dict(boxstyle='round', facecolor='white', alpha=0.5)

# place a text box in upper left in axes coords
ax.text(0.99, 0.01, textstr, transform=ax.transAxes, fontsize=10,
        verticalalignment='bottom', ha='right', bbox=props)
# [for d in dependencies]
plt.tight_layout()
plt.savefig('dependencies')
    
print(predicted_output['total_mass'])
print(out['total_mass'])
r2 = get_r2(out['total_mass'], predicted_output['total_mass'])
print(r2)



if 1:
    plt.figure()
    plt.plot(out['total_mass'], predicted_output['total_mass'],'.')
    plt.plot([0, 3e6], [0, 3e6])
    plt.title(f'Total Mass (r^2={r2:.3f})')
    plt.xlabel('Simulation mass [kg]')
    plt.ylabel('Mass from surrogate model [kg]')
    plt.grid()
    plt.tight_layout()
    plt.savefig('sim_vs_qls_mass')
if 0:
    for in_name in in_cols:
        plt.figure()
        plt.scatter(out['total_mass'], predicted_output['total_mass'], s=10, c=inp[in_name])
        plt.plot([0, 3e6], [0, 3e6])
        plt.xlabel('Simulation mass [kg]')
        plt.ylabel('Mass from surrogate model [kg]')
        plt.title(f'Sensitivity to {in_name}')
        plt.colorbar()
        plt.tight_layout()
if 0:
    plot_df = df.copy()
    plot_df = plot_df[plot_df['SP'] == 300]
    plot_df = plot_df[plot_df['WaterDepth'] == 20.5]
    plot_df = plot_df[plot_df['WaveHeight'] == 3.5]
    for in_col in in_cols:
        if in_col=='RP':
            for col2 in in_cols:
                if not col2==in_col:
                    plt.figure()
                    plt.scatter(plot_df[in_col], plot_df['total_mass'], s=10, c=plot_df[col2])
                    plt.xlabel(f'{in_col}')
                    plt.ylabel('Simulation mass')
                    plt.title(f'Sensitivity to {col2}')
                    plt.colorbar()
                    plt.tight_layout()

if 0:
    i_output_channel = 2
    RP = np.sort(np.asarray(df.RP.unique()))
    SP = 300
    WaterDepth = 20.5
    WaveHeight = 3.5
    HTrans = []
    HHub_Ratio = []
    WavePeriod = []
    WindSpeed = []
    D = []
    plt.figure()
    for rp in RP:
        plot_df = df.copy()
        plot_df = plot_df[plot_df['RP'] == rp]
        plot_df = plot_df[plot_df['SP'] == SP]
        plot_df = plot_df[plot_df['WaterDepth'] == WaterDepth]
        plot_df = plot_df[plot_df['WaveHeight'] == WaveHeight]
        HTrans.append(plot_df.HTrans.mean())
        HHub_Ratio.append(plot_df.HHub_Ratio.mean())
        WavePeriod.append(plot_df.WavePeriod.mean())
        WindSpeed.append(plot_df.WindSpeed.mean())
        D.append(plot_df.D.mean())
    output_channel_name = out_cols[i_output_channel]
    input_scaled = input_scaler.transform(np.array([RP,
                                                        SP*np.ones_like(RP),
                                                        D,
                                                        HTrans,
                                                        HHub_Ratio,
                                                        WaterDepth*np.ones_like(RP),
                                                        WaveHeight*np.ones_like(RP),
                                                        WavePeriod,
                                                        WindSpeed]).T)
    scaled_output = models[i_output_channel].getMetaModel()(input_scaled)
    output = output_scalers[output_channel_name].inverse_transform(scaled_output).ravel()
    plot_df = df.copy()
    plot_df = plot_df[plot_df['SP'] == SP]
    plot_df = plot_df[plot_df['WaterDepth'] == WaterDepth]
    plot_df = plot_df[plot_df['WaveHeight'] == WaveHeight]
    plt.scatter(plot_df['RP'], plot_df['total_mass'], label='data')
    plt.plot(RP, output, 'red', label='predicted')
    plt.grid()
    plt.legend()
    plt.xlabel('Rated Power [MW]')
    plt.ylabel('Total Mass [kg]')
    plt.title('Predicted total mass as function of rated power')

if 0:
    inps = np.array([ 10,200,13,0.7,20,3,5.00,8])
    inps_scaled = input_scaler.transform(inps.reshape(1, -1))
    scaled_output = models[2].getMetaModel()(inps_scaled)
    output = output_scalers['total_mass'].inverse_transform(scaled_output).ravel()
    print(float(output))                
    scaled_output = models[1].getMetaModel()(inps_scaled)
    output = output_scalers['tower_mass'].inverse_transform(scaled_output).ravel()
    print(float(output))                
    scaled_output = models[0].getMetaModel()(inps_scaled)
    output = output_scalers['monopile_mass'].inverse_transform(scaled_output).ravel()
    print(float(output))                
